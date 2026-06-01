"""CRYO Report Tools — compile_report, generate_excel, generate_chart."""

import base64
import json
import logging
import os
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

from tools.registry import registry

logger = logging.getLogger("cryo.reports")

API_PATH = Path(__file__).resolve().parent.parent.parent / "api"
if str(API_PATH.parent) not in sys.path:
    sys.path.insert(0, str(API_PATH.parent))

def _get_data_dir() -> Path:
    """Resolve data directory from env or project root."""
    env_dir = os.getenv("CRYO_DATA_DIR", "").strip()
    if env_dir:
        p = Path(env_dir)
        if p.is_absolute():
            return p
        # Relative path — resolve from project root
        return (Path(__file__).parent.parent.parent / env_dir).resolve()
    # Fallback: project_root/cryo-data
    return (Path(__file__).parent.parent.parent / "cryo-data").resolve()

DATA_DIR = _get_data_dir()
MAX_CONVERSATIONS_PER_USER = int(os.getenv("CRYO_MAX_CONVERSATIONS_PER_USER", "50"))


def _get_output_dir() -> Path:
    """Get the output directory for the current user/conversation.

    Structure: /cryo-data/users/{user_id}/conversations/{conversation_id}/reports/
    Falls back to /cryo-data/reports/ if no user context.
    """
    user_id = os.getenv("CRYO_USER_ID", "")
    convo_id = os.getenv("CRYO_CONVERSATION_ID", "")

    if user_id and convo_id:
        user_dir = DATA_DIR / "users" / user_id
        convo_dir = user_dir / "conversations" / convo_id / "reports"
        convo_dir.mkdir(parents=True, exist_ok=True)

        # Enforce max 50 conversations per user — delete oldest
        _cleanup_old_conversations(user_dir / "conversations")

        return convo_dir

    fallback = DATA_DIR / "reports"
    fallback.mkdir(parents=True, exist_ok=True)
    return fallback


def _get_sources_dir() -> Path:
    user_id = os.getenv("CRYO_USER_ID", "")
    convo_id = os.getenv("CRYO_CONVERSATION_ID", "")

    if user_id and convo_id:
        d = DATA_DIR / "users" / user_id / "conversations" / convo_id / "sources"
        d.mkdir(parents=True, exist_ok=True)
        return d

    fallback = DATA_DIR / "sources"
    fallback.mkdir(parents=True, exist_ok=True)
    return fallback


def _cleanup_old_conversations(conversations_dir: Path):
    """Keep only the newest MAX_CONVERSATIONS_PER_USER conversation dirs."""
    try:
        if not conversations_dir.exists():
            return
        convos = sorted(
            [d for d in conversations_dir.iterdir() if d.is_dir()],
            key=lambda d: d.stat().st_mtime,
            reverse=True,
        )
        for old_dir in convos[MAX_CONVERSATIONS_PER_USER:]:
            import shutil
            shutil.rmtree(old_dir, ignore_errors=True)
            logger.info("Cleaned old conversation dir: %s", old_dir.name[:12])
    except Exception as e:
        logger.warning("Conversation cleanup failed: %s", e)


def _gen_filename(prefix: str, ext: str) -> tuple[str, Path]:
    fid = uuid.uuid4().hex[:12]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{prefix}_{ts}_{fid}.{ext}"
    return name, _get_output_dir() / name


# ═══════════════════════════════════════════════════════════
# compile_report — markdown → interactive HTML report
# ═══════════════════════════════════════════════════════════

_last_report = {"title": "", "content": "", "citations": [], "filename": ""}


def _compile_report(args: dict, **kw) -> str:
    global _last_report
    title = args.get("title", "CRYO Research Report")
    content = args.get("content", "")
    citations_raw = args.get("citations", [])

    if not content:
        return json.dumps({"error": "Provide 'content' — full research in markdown"})

    logger.info("compile_report: title=%r len=%d citations=%d", title, len(content), len(citations_raw))

    try:
        # Parse markdown into sections
        sections = []
        summary = ""
        current_heading = None
        current_lines = []

        for line in content.split("\n"):
            h2 = re.match(r"^##\s+(.+)", line)
            if h2:
                if current_heading and current_lines:
                    sections.append({"heading": current_heading, "content": "\n".join(current_lines)})
                elif not current_heading and current_lines:
                    summary = "\n".join(current_lines).strip()
                current_heading = h2.group(1).strip()
                current_lines = []
            else:
                current_lines.append(line)

        if current_heading and current_lines:
            sections.append({"heading": current_heading, "content": "\n".join(current_lines)})
        elif current_lines and not sections:
            sections.append({"heading": "Research Findings", "content": "\n".join(current_lines)})

        if not summary and sections:
            first = sections[0].get("content", "")
            paras = [p.strip() for p in first.split("\n\n") if p.strip()]
            summary = paras[0] if paras else ""

        # Format citations
        citations = []
        for i, c in enumerate(citations_raw):
            if isinstance(c, str):
                citations.append({"id": i + 1, "text": c, "url": "", "doi": ""})
            elif isinstance(c, dict):
                citations.append({
                    "id": c.get("id", i + 1),
                    "text": c.get("text", c.get("formatted", "")),
                    "url": c.get("url", ""),
                    "doi": c.get("doi", ""),
                })

        # Render via report engine
        from api.services.report_engine import generate_report
        result = generate_report({
            "title": title,
            "summary": summary,
            "sections": sections,
            "citations": citations,
            "metadata": {
                "data_sources": ["PubMed", "OpenTargets", "ChEMBL", "ClinVar", "CrossRef"],
                "verification_status": "moderate",
            },
        })

        # Save raw source for editing later
        report_filename = result.get("filename", "")
        source_name = report_filename.replace(".html", ".json") if report_filename else ""
        if source_name:
            source_path = _get_sources_dir() / source_name
            source_path.write_text(json.dumps({
                "title": title,
                "content": content,
                "citations": citations_raw,
                "filename": report_filename,
            }, ensure_ascii=False, indent=2))
            logger.info("Source saved: %s", source_name)

        _last_report["title"] = title
        _last_report["content"] = content
        _last_report["citations"] = citations_raw
        _last_report["filename"] = report_filename

        return json.dumps(result)

    except Exception as e:
        logger.error("compile_report failed: %s", e, exc_info=True)
        # Minimal fallback — save raw content as HTML
        filename, filepath = _gen_filename("report", "html")
        filepath.write_text(f"<html><body><h1>{title}</h1><pre>{content}</pre></body></html>")
        return json.dumps({
            "status": "success", "filename": filename,
            "download_url": f"/api/reports/{filename}",
            "size_bytes": filepath.stat().st_size, "engine": "fallback",
        })


# ═══════════════════════════════════════════════════════════
# get_last_report — retrieve raw markdown of last report for editing
# ═══════════════════════════════════════════════════════════

def _get_last_report(args: dict, **kw) -> str:
    report_id = args.get("report_id", "")

    # Try memory first
    if _last_report["content"] and not report_id:
        return json.dumps({
            "title": _last_report["title"],
            "content": _last_report["content"],
            "citations": _last_report["citations"],
            "filename": _last_report["filename"],
            "content_length": len(_last_report["content"]),
            "instructions": "Modify the content as requested, then call compile_report with the updated content and citations.",
        })

    # Try loading from disk (by report_id or latest)
    try:
        sources_dir = _get_sources_dir()
        if report_id:
            source_files = list(sources_dir.glob(f"*{report_id}*.json"))
        else:
            source_files = sorted(sources_dir.glob("report_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)

        if source_files:
            data = json.loads(source_files[0].read_text())
            return json.dumps({
                "title": data.get("title", ""),
                "content": data.get("content", ""),
                "citations": data.get("citations", []),
                "filename": data.get("filename", ""),
                "content_length": len(data.get("content", "")),
                "source_file": source_files[0].name,
                "instructions": "Modify the content as requested, then call compile_report with the updated content and citations.",
            })
    except Exception as e:
        logger.error("Failed to load report source: %s", e)

    return json.dumps({"error": "No report found. Generate a report first with /report."})


registry.register(
    name="get_last_report",
    toolset="cryo_reports",
    schema={
        "name": "get_last_report",
        "description": "Retrieve the raw markdown content of the last generated report. Use this when the user asks to modify, edit, expand, or add to an existing report. After getting the content, modify it and call compile_report again with the updated content.",
        "parameters": {"type": "object", "properties": {
            "report_id": {"type": "string", "description": "Optional report filename to retrieve a specific report. Leave empty for the most recent."},
        }},
    },
    handler=_get_last_report,
    check_fn=lambda: True,
    emoji="📄",
)


COMPILE_SCHEMA = {
    "name": "compile_report",
    "description": (
        "Convert research into an interactive HTML report with charts, diagrams, tables, and callouts. "
        "Write 2000+ words in markdown with :::chart, :::diagram, :::callout, :::progress, :::timeline blocks. "
        "Use ## for sections, | for tables, **bold** for metrics, [1] for citations."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "title": {"type": "string", "description": "Report title"},
            "content": {
                "type": "string",
                "description": (
                    "Full research in markdown. Use :::chart, :::diagram, :::callout, :::progress, :::timeline blocks. "
                    "Use ## for sections, | col | for tables. Write 2000+ words with real data."
                ),
            },
            "citations": {
                "type": "array",
                "description": "APA citations",
                "items": {
                    "type": "object",
                    "properties": {
                        "text": {"type": "string"},
                        "url": {"type": "string"},
                        "doi": {"type": "string"},
                    },
                },
            },
        },
        "required": ["title", "content"],
    },
}

registry.register(name="compile_report", toolset="cryo_reports", schema=COMPILE_SCHEMA,
                  handler=_compile_report, check_fn=lambda: True, emoji="📋")
registry.register(name="generate_pdf", toolset="cryo_reports", schema=COMPILE_SCHEMA,
                  handler=_compile_report, check_fn=lambda: True, emoji="📑")


# ═══════════════════════════════════════════════════════════
# generate_excel
# ═══════════════════════════════════════════════════════════

def _generate_excel(args: dict, **kw) -> str:
    sheets = args.get("sheets", [])
    if not sheets:
        return json.dumps({"error": "Provide 'sheets' list"})

    logger.info("generate_excel: sheets=%d", len(sheets))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        filename, filepath = _gen_filename("data", "xlsx")
        wb = Workbook()

        header_font = Font(name="Calibri", size=11, bold=True, color="00E5C7")
        header_fill = PatternFill(start_color="0F1419", end_color="0F1419", fill_type="solid")
        cell_font = Font(name="Calibri", size=10, color="4B5563")
        cell_border = Border(bottom=Side(style="thin", color="E5E7EB"))

        for i, sheet_data in enumerate(sheets):
            name = sheet_data.get("name", f"Sheet{i+1}")[:31]
            data = sheet_data.get("data", [])

            ws = wb.active if i == 0 else wb.create_sheet(name)
            if i == 0:
                ws.title = name

            if not data or not isinstance(data[0], dict):
                continue

            headers = list(data[0].keys())

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row_data in enumerate(data[:10000], 2):
                for col, h in enumerate(headers, 1):
                    val = row_data.get(h, "")
                    cell = ws.cell(row=row_idx, column=col, value=val if isinstance(val, (int, float)) else str(val))
                    cell.font = cell_font
                    cell.border = cell_border

            for col in range(1, len(headers) + 1):
                max_len = max(
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, min(len(data) + 2, 50))
                )
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

            ws.freeze_panes = "A2"

        wb.save(str(filepath))
        logger.info("Excel generated: %s (%d bytes)", filename, filepath.stat().st_size)
        return json.dumps({
            "status": "success", "filename": filename,
            "download_url": f"/api/reports/{filename}",
            "size_bytes": filepath.stat().st_size,
        })

    except Exception as e:
        logger.error("Excel failed: %s", e, exc_info=True)
        return json.dumps({"error": f"Excel generation failed: {e}"})


registry.register(
    name="generate_excel", toolset="cryo_reports",
    schema={
        "name": "generate_excel",
        "description": "Generate Excel spreadsheet with multiple sheets.",
        "parameters": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "sheets": {"type": "array", "items": {"type": "object", "properties": {
                    "name": {"type": "string"},
                    "data": {"type": "array", "items": {"type": "object"}},
                }}},
            },
            "required": ["sheets"],
        },
    },
    handler=_generate_excel, check_fn=lambda: True, emoji="📊",
)


# ═══════════════════════════════════════════════════════════
# generate_chart (standalone PNG)
# ═══════════════════════════════════════════════════════════

def _generate_chart(args: dict, **kw) -> str:
    chart_type = args.get("chart_type", "bar")
    title = args.get("title", "Chart")
    data = args.get("data", {})

    if not data:
        return json.dumps({"error": "Provide 'data' with labels and values"})

    logger.info("generate_chart: type=%s title=%r", chart_type, title)

    try:
        from api.services.report_engine import render_chart
        b64 = render_chart({
            "type": chart_type, "title": title,
            "labels": data.get("labels", []),
            "values": data.get("values", []),
            "x_label": args.get("x_label", ""),
            "y_label": args.get("y_label", ""),
        })
        if not b64:
            return json.dumps({"error": "Chart render returned empty"})

        filename, filepath = _gen_filename("chart", "png")
        filepath.write_bytes(base64.b64decode(b64))
        logger.info("Chart saved: %s (%d bytes)", filename, filepath.stat().st_size)
        return json.dumps({
            "status": "success", "filename": filename,
            "download_url": f"/api/reports/{filename}",
            "size_bytes": filepath.stat().st_size,
        })

    except Exception as e:
        logger.error("Chart failed: %s", e, exc_info=True)
        return json.dumps({"error": f"Chart generation failed: {e}"})


registry.register(
    name="generate_chart", toolset="cryo_reports",
    schema={
        "name": "generate_chart",
        "description": "Generate standalone chart image (PNG).",
        "parameters": {
            "type": "object",
            "properties": {
                "chart_type": {"type": "string", "enum": ["bar", "horizontal_bar", "pie", "donut", "line", "scatter"]},
                "title": {"type": "string"},
                "data": {"type": "object", "description": "{labels: [...], values: [...]}"},
                "x_label": {"type": "string"},
                "y_label": {"type": "string"},
            },
            "required": ["chart_type", "title", "data"],
        },
    },
    handler=_generate_chart, check_fn=lambda: True, emoji="📈",
)
